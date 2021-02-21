#!/usr/bin/python3
import matplotlib.pyplot as plt
import openpyxl as xl

wb = xl.load_workbook('mem.xlsx')
sheet = wb.active
memList = []
for i in range(1, sheet.max_row):
    #print(sheet.cell(i, 1).value[0: -2])
    memList.append(int(sheet.cell(i, 1).value[0: -2]))

plt.plot(memList)
plt.show()
